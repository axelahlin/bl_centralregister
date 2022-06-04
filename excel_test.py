import os
from openpyxl import load_workbook
from datetime import date
print(os.getcwd())


wb = load_workbook("/home/axel/Downloads/vb.xlsx")
print('Loaded following sheets: ', wb.sheetnames)



q_vb = wb['Quratel + Valberedning']
seniors_osv = wb['Seniors osv.']
musik_klubb = wb['Musik och klubb']
planering = wb['Planeringsutskott']
cafe_other = wb['Café och övriga']
mat = wb['Matutskott']
info = wb['Förklaring och stadga']

year_sheetlimit = (date.today().year - 2010)*2 + 1
terms = q_vb['A']

t = terms[1:year_sheetlimit]




search_name = input("Enter your value: ")
print("\n")



#procedur för räknar ut totala antalet mandat
occurences = 0
occupied_terms = [False] * len(terms)


for s in wb.worksheets:
    for row in s:
        for cell in row:
            if cell.value != None and search_name == cell.value: #in substring för alla, == för icke vakanser
                r = cell.row-1
                c = cell.column-1

                if occupied_terms[r] != False:
                    print(terms[r].value,s[1][c].value)
                
                else:
                    occurences+=1
                #print(r,c)
                #print(terms[r].value,s[1][c].value) #printa titel, -1 för faktisk

print(" Totalt", occurences, " HELA terminer med så här många mandat/poster")





# for i in range(1,year_sheetlimit):
#     print(q_vb['A'+str(i)].value, q_vb['F'+str(i)].value)






